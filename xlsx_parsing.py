from openpyxl import load_workbook
import csv

source_workbook = load_workbook(filename="SBA Table of Size Standards_Effective Aug 19, 2019.xlsx")

print(source_workbook.sheetnames)
 

for sheet_num, sheet_name in enumerate(source_workbook.sheetnames):
    if sheet_name ==  'table_of_size_standards-all':

        current_sheet = source_workbook[sheet_name]
        print(sheet_name)

        output_file = open("xlsx_" + sheet_name + ".csv", 'w')
        output_writer = csv.writer(output_file, dialect='unix')

        for row in current_sheet.iter_rows():

            row_cells = []

            for cell in row:
                print(cell, cell.value)
                row_cells.append(str(cell.value))

            output_writer.writerow(row_cells)
    else:
        pass

output_file.close()
